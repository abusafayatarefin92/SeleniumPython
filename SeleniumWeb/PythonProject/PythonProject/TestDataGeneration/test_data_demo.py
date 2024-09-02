from faker import Faker
from openpyxl import Workbook

work_book = Workbook()
work_sheet = work_book.active

fake_data = Faker()
# print(fake_data.name())
# print(fake_data.email())
# print(fake_data.address())

for i in range(1, 11):
    for j in range(1, 4):
        work_sheet.cell(i, 1).value = fake_data.name()
        work_sheet.cell(i, 2).value = fake_data.email()
        work_sheet.cell(i, 3).value = fake_data.address()
    # print(fake_data.name())

work_book.save('testdata.xlsx')
