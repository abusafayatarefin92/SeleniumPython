from openpyxl import Workbook, load_workbook

work_book = load_workbook(filename= 'demoexcel.xlsx')
# worksheet = work_book.active
work_sheet = work_book['DemoSheet']

print(work_sheet['A4'].value)
print(work_book['DemoSheet']['A3'].value)
print(work_sheet.cell(2, 2).value)

row_count = work_sheet.max_row
column_count = work_sheet.max_column

print(row_count, column_count)

for i in range(1, row_count + 1):
    for j in range(1, column_count + 1):
        print(work_sheet.cell(i, j).value, end= ' ')
    print('\n')
